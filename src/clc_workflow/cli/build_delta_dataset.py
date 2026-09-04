#!/usr/bin/env python3
"""
Build a mixed-type dpdata dataset of measured oxygen capacity for property training.

    clc delta config_feco.yaml --xlsx <experiment.xlsx> [<more.xlsx> ...] \
           [--sheet S | --sheet S1 S2 ...] \
           [--source opt|md_avg] [--out delta_dataset] \
           [--valid-frac F | --kfold K] [--dry-run]

Several spreadsheets are read and STACKED into one dataset -- separate measurement
campaigns feeding one training set.  Every row keeps a `source` column naming the file it
came from, so `source` + `block` + `xlsx_row` still identifies one cell; without it the
row numbers of two workbooks would collide silently.  A (compound, window, T) measured in
more than one file is reported rather than quietly duplicated, since a repeated label is
either a genuine replicate or the same file passed twice.

One frame per (structure, measurement).  The structure supplies coordinates and cell;
the experiment supplies the label and the conditions it was measured at:

    fparam   = [P_O2_high (atm), P_O2_low (atm), T]      per frame

The spreadsheet's own temperature unit is read from its cells ('400 °C', '673 K') and
converted to whatever --t-unit asks for.  An unlabelled sheet is assumed Celsius and says
so loudly, because a bare 700 is an ordinary measurement in either unit and guessing it
wrong shifts the whole temperature axis by a freezing point without any symptom.
    delta    = [dd]                                       per frame, the label

WHY BOTH PRESSURES.  dd is not a property of a state, it is the change in oxygen content
BETWEEN two states: dd = delta(P_low, T) - delta(P_high, T).  The spreadsheet holds two
blocks measured over different windows -- 0.8 atm -> 3e-5 atm and 0.2 atm -> 1e-3 atm --
and they share temperatures.  Storing a single pressure would make those rows look like
contradictory labels for the same conditions, so the window is carried in full.

EVERY SET IS A FRAME.  The n_sets SQS realisations of one composition are independent
samples of the same disordered material, so each contributes its own frame with the same
label.  That is the intended use of the sets, not double counting: the label belongs to
the composition, and the realisations are what tell the model which features of a
particular decoration are irrelevant to it.

The structures are identical across the measurements of one compound -- only fparam
differs -- which is exactly the shape a property model with frame parameters expects.

HOW IT IS SPLIT.  One held-out set by default (--valid-frac), or K rotating ones
(--kfold K) so every compound serves as validation exactly once -- `clc kfold` does the
same to a dataset that already exists.  Either way the split is
by COMPOUND, never by frame -- see the comment above the split for why.  The K folds are
written once each, as disjoint directories; a fold's training set is the other K-1
directories, listed for you in folds.json, so nothing is duplicated on disk.

Output is deepmd/npy/mixed: one directory per atom count, with real_atom_types.npy
carrying the per-frame species.  Mixed type is required here rather than convenient,
because A-site- and B-site-deficient cells have fewer atoms than the stoichiometric ones
and a plain deepmd/npy system cannot hold both.

`delta.npy` is written beside coord.npy.  The name has to match `property_name` in the
deepmd input.json; change both together with --label-name.
"""
import argparse
import json
import os
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd

from clc_workflow.clc_config import load_config, MANIFEST_NAME
from clc_workflow.kfold import assign_folds, fold_manifest
# source name -> (filenames to try, in order)
SOURCES = {
    "md_avg": ["POSCAR_md_avg", "POSCAR_md_final"],   # legacy name still read
    "opt": ["optimized_POSCAR"],
}
A_SITE = ("Sr", "Ca", "Ba", "La", "K", "Y", "Ce", "Sm")
B_SITE = ("Fe", "Co", "Mn", "Ti", "Ni", "Cr", "Mo", "Mg")
# Zero-width characters: the spreadsheet's formulas are full of U+200B, which is why
# "Sr0.75Ba0.25FeO3" from the sheet does not compare equal to the same string typed here.
_INVISIBLE = dict.fromkeys(map(ord, "​‌‍﻿\xa0"), None)
MATCH_TOL = 0.02          # site-fraction tolerance; the rounded A-excess rows need ~0.005


# ---------------------------------------------------------------------------- formulas

def clean_formula(text):
    """Strip the invisible characters and normalise the minus sign and the O3-d tail."""
    s = str(text).translate(_INVISIBLE).strip()
    s = s.replace("−", "-").replace("–", "-").replace(" ", "")
    s = re.sub(r"O3-?[δd]?$", "", s)      # drop the O3-delta tail; delta is the label
    return s


def parse_formula(text):
    """
    'Sr0.75Ca0.2Fe0.875Co0.125' -> {'Sr': .75, 'Ca': .2, 'Fe': .875, 'Co': .125}

    Handles one level of parentheses with a multiplier, which is how the sheet writes
    (Sr0.75Ba0.25)0.95FeO3 -- a 5% A-site deficiency spread over both A species.
    """
    s = clean_formula(text)
    out = {}

    def add(el, amt):
        out[el] = out.get(el, 0.0) + amt

    while "(" in s:
        m = re.search(r"\(([^()]*)\)(\d*\.?\d*)", s)
        if not m:
            raise ValueError(f"unbalanced parentheses in {text!r}")
        mult = float(m.group(2)) if m.group(2) else 1.0
        for el, amt in re.findall(r"([A-Z][a-z]?)(\d*\.?\d*)", m.group(1)):
            add(el, (float(amt) if amt else 1.0) * mult)
        s = s[:m.start()] + s[m.end():]

    for el, amt in re.findall(r"([A-Z][a-z]?)(\d*\.?\d*)", s):
        add(el, float(amt) if amt else 1.0)
    if not out:
        raise ValueError(f"no elements parsed from {text!r}")
    return out


def site_fractions(counts):
    """
    Element amounts (per formula unit) -> fractions of the A and B SITES.

    A perovskite has as many A sites as B sites, so the sublattice that is full sets the
    scale.  A/B = 1.05 means the A sublattice is full and 5% of the B sites are empty,
    which is how Sr0.75Ca0.3Fe...O3 has to be read -- see config_feco.yaml.
    """
    a = {el: v for el, v in counts.items() if el in A_SITE}
    b = {el: v for el, v in counts.items() if el in B_SITE}
    if not a or not b:
        raise ValueError(f"could not split {counts} into A and B sites")
    scale = max(sum(a.values()), sum(b.values()))
    return ({el: v / scale for el, v in a.items()},
            {el: v / scale for el, v in b.items()})


# --------------------------------------------------------------------------- the sheet

# A temperature-row cell: a number, optionally carrying its unit as text.
# \u2103 is the single-character DEGREE CELSIUS, \u212a the KELVIN SIGN -- both turn up
# in spreadsheets that were typed rather than computed.
_TEMP_CELL = re.compile(
    r"^\s*([-+]?\d*\.?\d+)\s*(?:°\s*)?([CcKk]|\u2103|\u212a)?\s*$")
_UNIT_OF = {"c": "C", "\u2103": "C", "k": "K", "\u212a": "K"}


def parse_temperature(v):
    """
    One cell of the temperature row -> (value, unit) with unit "C", "K" or None.

    The columns used to be bare numbers and `float(cell)` was enough.  SCFM260817.xlsx
    writes them as text instead -- '400 °C' -- so every cell failed the isinstance check
    and the block was reported as having no temperature row at all.  Accept both, and
    strip the zero-width characters the sheet is full of (see _INVISIBLE).

    float() rather than isinstance also fixes a latent case: a column of whole degrees
    comes back as numpy.int64, which is NOT an instance of int.

    The unit is now REPORTED rather than thrown away.  It used to be matched purely so
    that '400 °C' would parse, and the sheet was then assumed Celsius no matter what it
    said -- which means a Kelvin sheet had 273.15 added to temperatures that were already
    Kelvin, silently, and the model trained on a temperature axis shifted by a whole
    freezing point.  A bare number still yields None, because a number alone genuinely
    does not say: 700 is a plausible measurement in either unit.
    """
    if isinstance(v, str):
        m = _TEMP_CELL.match(v.translate(_INVISIBLE))
        if not m:
            return None, None
        return float(m.group(1)), _UNIT_OF.get((m.group(2) or "").lower())
    if v is None or pd.isna(v):
        return None, None
    try:
        return float(v), None
    except (TypeError, ValueError):
        return None, None


def _sheet_arg(v):
    """argparse hands back strings; a bare number means a sheet INDEX, not a name."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def read_measurements(xlsx, sheet=0):
    """
    Every (compound, pressure window, temperature) cell of the sheet as one row.

    The layout is a stack of blocks: a header naming the pressure window, a row of
    temperatures, then one row per compound.  Blocks are found by their header rather
    than by fixed row numbers, so adding a block or a temperature column does not need a
    code change.  A compound whose whole row is red is marked `flagged` -- the sheet's own
    note says those measurements came out near zero, negative, or rising anomalously at
    low pO2, i.e. they are not labels you want a model to fit.
    """
    raw = pd.read_excel(xlsx, sheet_name=sheet, header=None)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx)
        ws = wb[wb.sheetnames[sheet]] if isinstance(sheet, int) else wb[sheet]
        red_rows = set()
        for row in ws.iter_rows():
            for c in row:
                rgb = getattr(getattr(c.font, "color", None), "rgb", None)
                if isinstance(rgb, str) and rgb.upper() == "FFFF0000":
                    red_rows.add(c.row - 1)          # openpyxl is 1-based
                    break
    except Exception as e:                            # formatting is a hint, not the data
        print(f"[warn] could not read cell colours ({e}); nothing will be flagged")
        red_rows = set()

    blocks = []
    for r in range(len(raw)):
        text = " ".join(str(v) for v in raw.iloc[r].tolist() if pd.notna(v))
        m = re.search(r"PO2[:：]\s*([\d.eE+-]+)\s*atm\s*[-–—>→]+\s*([\d.eE+-]+)\s*atm", text)
        if m:
            blocks.append((r, float(m.group(1)), float(m.group(2))))
    if not blocks:
        raise SystemExit(f"[ERROR] {xlsx}: found no 'PO2: <hi> atm -> <lo> atm' block header")

    rows = []
    for i, (hdr, p_hi, p_lo) in enumerate(blocks):
        end = blocks[i + 1][0] if i + 1 < len(blocks) else len(raw)
        temps = {}
        for c in range(raw.shape[1]):
            t, unit = parse_temperature(raw.iat[hdr + 1, c])
            if t is not None:
                temps[c] = (t, unit)
        if not temps:
            raise SystemExit(f"[ERROR] {xlsx}: no temperature row under the block header "
                             f"at spreadsheet row {hdr + 2}.  Expected numbers, or text "
                             f"carrying the unit ('400 °C'), one per temperature column.")
        for r in range(hdr + 2, end):
            # The first cell of the row that parses into both an A and a B species.
            # Pattern-matching the text instead would drop SrFeO3-d, whose leading
            # element carries no subscript, and it did.
            label = None
            for c in range(min(5, raw.shape[1])):
                v = raw.iat[r, c]
                if not isinstance(v, str):
                    continue
                try:
                    counts = parse_formula(v)
                except ValueError:
                    continue
                if set(counts) & set(A_SITE) and set(counts) & set(B_SITE):
                    label = v
                    break
            if label is None:
                continue
            for c, (t, unit) in temps.items():
                v = raw.iat[r, c]
                if not isinstance(v, (int, float)) or pd.isna(v):
                    continue
                rows.append({"formula_raw": str(label), "formula": clean_formula(label),
                             "p_high_atm": p_hi, "p_low_atm": p_lo,
                             "t_sheet": t, "t_unit_sheet": unit,
                             "delta": float(v), "flagged": r in red_rows,
                             "source": os.path.basename(str(xlsx)),
                             "block": i + 1, "xlsx_row": r + 1})
    return pd.DataFrame(rows)


# the columns that make one measurement the same measurement, for duplicate reporting
_MEAS_KEY = ["formula", "p_high_atm", "p_low_atm", "t_sheet"]


def read_all_measurements(xlsxs, sheets):
    """
    Read every spreadsheet and stack them, keeping each row's origin.

    `sheets` is one sheet applied to all files, or one per file.  Reading them separately
    rather than concatenating the workbooks matters for the red-cell scan: `flagged` comes
    from cell formatting, which is per workbook.
    """
    if len(sheets) == 1:
        sheets = sheets * len(xlsxs)
    if len(sheets) != len(xlsxs):
        raise SystemExit(
            f"[ERROR] --sheet was given {len(sheets)} value(s) for {len(xlsxs)} "
            f"spreadsheet(s).\n"
            f"        Give one sheet to apply to all of them, or one per file in the "
            f"same order.")

    frames = []
    for xlsx, sheet in zip(xlsxs, sheets):
        if not os.path.isfile(xlsx):
            raise SystemExit(f"[ERROR] no such spreadsheet: {xlsx}")
        m = read_measurements(xlsx, sheet)
        if m.empty:
            print(f"[warn] {xlsx}: no measurements read; is --sheet right?")
        frames.append(m)

    seen = [f for f in frames if not f.empty]
    if not seen:
        raise SystemExit("[ERROR] no measurements in any spreadsheet")
    meas = pd.concat(seen, ignore_index=True)

    # A basename appearing twice would make `source` ambiguous, which defeats the point.
    bases = [os.path.basename(str(x)) for x in xlsxs]
    if len(set(bases)) != len(bases):
        raise SystemExit(
            f"[ERROR] two spreadsheets share a filename: "
            f"{', '.join(sorted({b for b in bases if bases.count(b) > 1}))}.\n"
            f"        `source` records the basename, so it could not tell the rows "
            f"apart.  Rename one, or copy them to distinct names first.")
    return meas


def resolve_sheet_unit(meas, declared):
    """
    What unit the spreadsheet's temperatures are in -> ("C"|"K", how it was decided).

    A label in the cell wins outright.  Labels that disagree with each other, or with an
    explicit --sheet-t-unit, are an error: the sheet is then saying two things and only
    the person who made it knows which is true.

    With no label at all the number alone is usually not enough -- 700 is a perfectly
    ordinary measurement in either unit -- so this does NOT guess.  The one case it can
    settle is a sheet whose temperatures are all below 273.15, which cannot be Kelvin for
    a solid-state measurement.  Everything else falls back to `declared`, and says so.
    """
    labels = {u for u in meas["t_unit_sheet"].dropna().unique()}
    if len(labels) > 1:
        where = (meas[meas["t_unit_sheet"].notna()]
                 .groupby(["source", "t_unit_sheet"])["t_sheet"]
                 .agg(lambda g: ", ".join(f"{v:g}" for v in sorted(set(g))[:4])))
        raise SystemExit(
            "[ERROR] the spreadsheet labels its temperatures with more than one unit:\n"
            + "\n".join(f"        {src}: {unit} at {vals}"
                         for (src, unit), vals in where.items())
            + "\n        Fix the sheet, or pass --sheet-t-unit to say which is right "
              "for all of it.")

    if labels:
        found = labels.pop()
        if declared not in (None, "auto") and declared != found:
            raise SystemExit(
                f"[ERROR] --sheet-t-unit {declared} contradicts the spreadsheet, which "
                f"labels its temperatures {found}.\n"
                f"        One of the two is wrong and this cannot tell which; correct "
                f"the sheet or drop the flag.")
        return found, f"read from the sheet's own labels"

    if declared not in (None, "auto"):
        return declared, "--sheet-t-unit"

    hottest = float(meas["t_sheet"].max())
    if hottest < 273.15:
        return "C", f"unlabelled, but the highest is {hottest:g} -- not Kelvin"
    return "C", "unlabelled; ASSUMED (see the warning above)"


def convert_temperature(values, src, dst):
    """Celsius <-> Kelvin, or a no-op when the units already agree."""
    if src == dst:
        return values
    return values + 273.15 if dst == "K" else values - 273.15


def report_duplicates(meas):
    """Measurements repeated across files -- a replicate, or the same file twice."""
    if meas["source"].nunique() < 2:
        return
    dup = meas[meas.duplicated(_MEAS_KEY, keep=False)]
    if dup.empty:
        return
    groups = dup.groupby(_MEAS_KEY)
    print(f"[warn] {groups.ngroups} measurement(s) appear in more than one spreadsheet; "
          f"all copies are kept as separate frames")
    for (formula, hi, lo, t), g in list(groups)[:5]:
        where = ", ".join(f"{r['source']}:row{r['xlsx_row']}={r['delta']:g}"
                          for _, r in g.iterrows())
        print(f"       {formula}  {hi:g}->{lo:g} atm  {t:g} C   {where}")
    if groups.ngroups > 5:
        print(f"       ... and {groups.ngroups - 5} more")


# ------------------------------------------------------------------------- the manifest

def manifest_site_fractions(row):
    """Realised A- and B-site fractions of a manifest row, from its integer site counts."""
    out = []
    for base, dop, frac, vac, n_sites in (
            (row["a_base"], row["a_dopant"], row["x"], row.get("a_vac", 0.0),
             row.get("n_A_sites", row["n_A"])),
            (row["b_base"], row["b_dopant"], row["y"], row.get("b_vac", 0.0),
             row.get("n_B_sites", row["n_B"]))):
        n_sites = int(n_sites)
        n_dop = int(round(float(frac) * n_sites))
        n_vac = int(round(float(vac or 0.0) * n_sites))
        d = {}
        if n_sites - n_dop - n_vac:
            d[base] = (n_sites - n_dop - n_vac) / n_sites
        if n_dop:
            d[dop] = n_dop / n_sites
        out.append(d)
    return out[0], out[1]


def match_formula(a_want, b_want, comps, tol=MATCH_TOL):
    """Composition -> the manifest's composition key, plus the worst site-fraction error."""
    hits = []
    for key, (a_have, b_have) in comps.items():
        err = 0.0
        for want, have in ((a_want, a_have), (b_want, b_have)):
            for el in set(want) | set(have):
                err = max(err, abs(want.get(el, 0.0) - have.get(el, 0.0)))
        if err <= tol:
            hits.append((err, key))
    hits.sort()
    return hits


# ------------------------------------------------------------------------------- output

def register_types(label_name):
    """Teach dpdata about fparam and the label so the deepmd writers emit both."""
    import dpdata
    from dpdata.data_type import Axis, DataType
    for cls in (dpdata.System, dpdata.LabeledSystem):
        for name in ("fparam", label_name):
            cls.register_data_type(
                DataType(name, np.ndarray, (Axis.NFRAMES, -1), required=False,
                         deepmd_name=name))


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("config")
    ap.add_argument("--xlsx", required=True, nargs="+",
                    help="the measured dd(P, T) spreadsheet(s); several are stacked into "
                         "one dataset and each row records which file it came from")
    ap.add_argument("--sheet", default=[0], nargs="+",
                    help="sheet name or index (default: the first).  One value applies "
                         "to every spreadsheet; or give one per --xlsx, in the same order")
    ap.add_argument("--source", choices=sorted(SOURCES), default="opt",
                    help="which structure to use: 'opt' = optimized_POSCAR (default), "
                         "'md_avg' = POSCAR_md_avg")
    ap.add_argument("--out", default="delta_dataset", help="output directory")
    ap.add_argument("--label-name", default="delta",
                    help="name of the label array; must equal property_name in the "
                         "deepmd input.json (default: delta)")
    ap.add_argument("--sets", default=None, help="comma list of set names to restrict to")
    ap.add_argument("--t-unit", choices=["K", "C"], default="K",
                    help="temperature unit written INTO fparam; the rest of this pipeline "
                         "works in K, so K is the default")
    ap.add_argument("--sheet-t-unit", choices=["C", "K", "auto"], default="auto",
                    help="the unit the SPREADSHEET is in.  'auto' (default) reads the "
                         "sheet's own labels ('400 °C', '673 K') and falls back to C when "
                         "there are none -- a bare number does not say which it is.  Set "
                         "it explicitly for an unlabelled Kelvin sheet, or the "
                         "temperatures get 273.15 added to them twice")
    ap.add_argument("--p-log10", action="store_true",
                    help="store log10(P/atm) instead of P/atm -- the pressures span four "
                         "decades, which a linear feature represents badly")
    ap.add_argument("--keep-flagged", action="store_true",
                    help="include the measurements the spreadsheet marks red as "
                         "unreliable (excluded by default)")
    ap.add_argument("--valid-frac", type=float, default=0.1,
                    help="fraction of FRAMES held out for validation (default 0.1). "
                         "The split is by compound: every frame of a compound -- all its "
                         "sets, pressures and temperatures -- lands on the same side. "
                         "0 writes a single undivided dataset")
    ap.add_argument("--seed", type=int, default=0,
                    help="seed for choosing which compounds are held out")
    ap.add_argument("--valid-compounds", default=None,
                    help="comma list of formulas to hold out, overriding the random "
                         "choice -- use it to pin a split across reruns")
    ap.add_argument("--kfold", type=int, default=0,
                    help="write K disjoint folds instead of one train/valid split, so "
                         "every compound is validation in exactly one fold.  Folds are "
                         "balanced on FRAMES and split by compound, like --valid-frac, "
                         "which this overrides.  Each fold is written once; folds.json "
                         "lists the train/valid systems for each of the K runs")
    ap.add_argument("--type-map", default=None,
                    help="comma list fixing the species order; default is derived from "
                         "the manifest.  It is written to type_map.raw either way")
    ap.add_argument("--set-size", type=int, default=5000, help="frames per set.* dir")
    ap.add_argument("--dry-run", action="store_true",
                    help="report the matching and the frame count; write nothing")
    args = ap.parse_args(argv)

    cfg = load_config(args.config)
    root = cfg["output_root"]
    man_path = os.path.join(root, MANIFEST_NAME)
    if not os.path.isfile(man_path):
        sys.exit(f"[ERROR] {man_path} not found -- run `clc sqs` first")
    man = pd.read_csv(man_path)
    if args.sets:
        keep = {s.strip() for s in args.sets.split(",") if s.strip()}
        man = man[man["set"].isin(keep)]
    if man.empty:
        sys.exit("[ERROR] no manifest rows selected")

    sheets = [_sheet_arg(v) for v in args.sheet]
    meas = read_all_measurements(args.xlsx, sheets)
    n_flagged = int(meas["flagged"].sum())
    per_file = meas.groupby("source").size()
    print(f"[*] spreadsheet  : "
          + ", ".join(f"{name} ({n} pts)" for name, n in per_file.items()))
    print(f"[*] measurements : {len(meas)} over {meas['formula'].nunique()} compound(s), "
          f"{n_flagged} flagged unreliable")

    sheet_unit, how = resolve_sheet_unit(meas, args.sheet_t_unit)
    if how.startswith("unlabelled; ASSUMED"):
        print(f"[!] the spreadsheet's temperatures carry no unit, and "
              f"{meas['t_sheet'].min():g}-{meas['t_sheet'].max():g} is a plausible range "
              f"in either C or K.\n"
              f"[!] ASSUMING CELSIUS.  If the sheet is in Kelvin, pass "
              f"--sheet-t-unit K -- otherwise every\n"
              f"[!] temperature in fparam is 273.15 too high and nothing will complain.\n"
              f"[!] Labelling one cell '{meas['t_sheet'].max():g} °C' or "
              f"'{meas['t_sheet'].max():g} K' settles it for good.")
    print(f"[*] sheet T unit : {sheet_unit}  ({how})")

    for (hi, lo), g in meas.groupby(["p_high_atm", "p_low_atm"]):
        print(f"[*]   window {hi:g} atm -> {lo:g} atm : {len(g)} point(s), "
              f"T = {', '.join(f'{t:g}' for t in sorted(g['t_sheet'].unique()))} "
              f"{sheet_unit}")
    report_duplicates(meas)
    if not args.keep_flagged and n_flagged:
        meas = meas[~meas["flagged"]].reset_index(drop=True)
        print(f"[*] dropped {n_flagged} flagged measurement(s); --keep-flagged to keep them")

    # one composition key per distinct (composition), pointing at all its structure dirs
    comps, dirs, comp_name = {}, {}, {}
    for _, row in man.iterrows():
        a, b = manifest_site_fractions(row)
        key = (row["a_base"], row["a_dopant"], round(float(row["x"]), 6),
               round(float(row.get("a_vac", 0.0) or 0.0), 6),
               row["b_base"], row["b_dopant"], round(float(row["y"]), 6),
               round(float(row.get("b_vac", 0.0) or 0.0), 6))
        comps[key] = (a, b)
        dirs.setdefault(key, []).append(row["path"])
        comp_name[key] = row.get("comp_dir", str(key))

    # ---- match every measured compound to a composition in the tree ----------------
    index, unmatched, ambiguous, missing_file = [], [], [], {}
    for formula, g in meas.groupby("formula"):
        try:
            a_want, b_want = site_fractions(parse_formula(formula))
        except ValueError as e:
            unmatched.append((formula, str(e)))
            continue
        hits = match_formula(a_want, b_want, comps)
        if not hits:
            unmatched.append((formula, "no composition within tolerance"))
            continue
        if len(hits) > 1 and abs(hits[0][0] - hits[1][0]) < 1e-9:
            ambiguous.append((formula, [k for _, k in hits]))
            continue
        err, key = hits[0]
        for path in dirs[key]:
            src = None
            for name in SOURCES[args.source]:
                cand = os.path.join(root, path, name)
                if os.path.isfile(cand):
                    src = cand
                    break
            if src is None:
                missing_file[path] = SOURCES[args.source][0]
                continue
            for _, m in g.iterrows():
                index.append({"path": path, "structure": os.path.relpath(src, root),
                              "formula": formula, "comp": comp_name[key],
                              "match_err": round(err, 6),
                              "p_high_atm": m["p_high_atm"], "p_low_atm": m["p_low_atm"],
                              "t_sheet": m["t_sheet"], "t_unit_sheet": sheet_unit,
                              "t_fparam": convert_temperature(
                                  float(m["t_sheet"]), sheet_unit, args.t_unit),
                              args.label_name: m["delta"],
                              "flagged": bool(m["flagged"]), "source": m["source"],
                              "block": m["block"], "xlsx_row": m["xlsx_row"]})

    idx = pd.DataFrame(index)
    matched = meas[meas["formula"].isin(idx["formula"].unique())] if len(idx) else meas.iloc[:0]
    print(f"\n[*] matched      : {idx['formula'].nunique() if len(idx) else 0}"
          f"/{meas['formula'].nunique()} compound(s) to the tree, "
          f"worst site-fraction error {idx['match_err'].max() if len(idx) else float('nan'):.4f}")
    print(f"[*] structures    : {idx['path'].nunique() if len(idx) else 0} directories "
          f"({args.source}), {len(idx)} frame(s)")
    for formula, why in unmatched:
        print(f"    [UNMATCHED] {formula}: {why}")
    for formula, keys in ambiguous:
        print(f"    [AMBIGUOUS] {formula}: matches {len(keys)} compositions equally well")
    if missing_file:
        shown = list(missing_file.items())[:5]
        for p, f in shown:
            print(f"    [NO {f}] {p}")
        if len(missing_file) > len(shown):
            print(f"    ... and {len(missing_file) - len(shown)} more without "
                  f"{SOURCES[args.source][0]}")
    if idx.empty:
        sys.exit("\n[ERROR] nothing to write.")

    if args.type_map:
        type_map = [s.strip() for s in args.type_map.split(",") if s.strip()]
    else:
        seen = []
        for col in ("a_base", "a_dopant", "b_base", "b_dopant"):
            for v in man[col].dropna().unique():
                if v not in seen:
                    seen.append(v)
        type_map = [e for e in seen if e in A_SITE] + [e for e in seen if e in B_SITE] + ["O"]
    print(f"[*] type_map     : {type_map}")

    # ---- train / valid, split BY COMPOUND ------------------------------------------
    # Never by frame.  Every frame of one compound shares a label that differs only
    # through fparam, and the n_sets realisations are the same material decorated
    # differently -- so a frame-wise split would put a compound's 500 K point in train and
    # its 600 K point in validation, and the validation score would measure interpolation
    # between two nearly identical rows rather than transfer to an unseen material.
    # Holding out whole compounds is the only split that answers the question being asked.
    if args.kfold:
        # ---- K rotating held-out sets --------------------------------------------
        # The same rule -- whole compounds move together -- applied K times, so
        # every compound is validation exactly once and no compound is ever in both
        # sides of the same fold.  This is group K-fold, not plain K-fold; plain
        # K-fold on frames would leak a compound's other temperatures and its other
        # SQS realisations into training and report a score that means nothing.
        if args.kfold < 2:
            sys.exit("[ERROR] --kfold needs at least 2 folds "
                     "(--kfold 0 or --valid-frac for a single split)")
        if args.valid_compounds:
            sys.exit("[ERROR] --valid-compounds pins one held-out set, --kfold rotates "
                     "through K of them.  Use one or the other")
        by_comp = idx.groupby("comp").size()
        if len(by_comp) < args.kfold:
            sys.exit(f"[ERROR] --kfold {args.kfold} needs at least {args.kfold} matched "
                     f"compounds; only {len(by_comp)} matched.  Every fold has to hold "
                     f"out a whole compound, so K cannot exceed the compound count")
        # Balanced on FRAMES, not on compound count -- a compound carries between 4 and
        # 7 measurements times n_sets frames, so dealing them round-robin would leave
        # folds of visibly different size and the K scores would not be comparable.
        # Shared with `clc kfold`, which does the same to an existing dataset.
        fold_of = assign_folds(by_comp.to_dict(), args.kfold, args.seed)
        idx["fold"] = idx["comp"].map(fold_of)
        idx["split"] = "fold_" + idx["fold"].astype(str)

        print(f"\n[*] split        : {args.kfold}-fold, held out by compound "
              f"({len(by_comp)} compound(s), {len(idx)} frame(s))")
        for k in range(args.kfold):
            g = idx[idx["fold"] == k]
            print(f"[*]   fold {k}: {len(g)} valid / {len(idx) - len(g)} train frame(s) "
                  f"({len(g) / len(idx):.1%}), {g['comp'].nunique()} compound(s)")
            for c in sorted(g["comp"].unique()):
                f = g.loc[g["comp"] == c, "formula"].iloc[0]
                print(f"[*]       {c}  ({f}, {int((idx['comp'] == c).sum())} frames)")
        splits = [(f"fold_{k}", idx[idx["fold"] == k]) for k in range(args.kfold)]
    else:
        idx["split"] = "train"
        if args.valid_frac > 0:
            by_comp = idx.groupby("comp").size()
            if args.valid_compounds:
                want = {s.strip() for s in args.valid_compounds.split(",") if s.strip()}
                held = sorted(idx.loc[idx["formula"].isin(want) | idx["comp"].isin(want),
                                      "comp"].unique())
                unknown = want - set(idx["formula"]) - set(idx["comp"])
                for u in unknown:
                    print(f"    [WARN] --valid-compounds {u!r} matched nothing")
            else:
                # Shuffle compounds, take them until the frame target is met.  Aiming at a
                # fraction of FRAMES rather than of compounds matters here: compounds carry
                # between 4 and 7 measurements each, so 10% of compounds is not 10% of data.
                rng = np.random.default_rng(args.seed)
                order = list(by_comp.index)
                rng.shuffle(order)
                target = args.valid_frac * len(idx)
                held, got = [], 0
                for c in order:
                    if got >= target:
                        break
                    held.append(c)
                    got += by_comp[c]
                # The compound that crosses the target usually overshoots it by more than
                # stopping short would undershoot -- with 4-7 measurements each, one compound
                # is several percent of the whole set.  Keep whichever side is closer.
                if len(held) > 1 and abs(got - by_comp[held[-1]] - target) < abs(got - target):
                    got -= by_comp[held.pop()]
                held = sorted(held)
            idx.loc[idx["comp"].isin(held), "split"] = "valid"

        n_valid = int((idx["split"] == "valid").sum())
        if args.valid_frac > 0:
            print(f"\n[*] split        : {len(idx) - n_valid} train / {n_valid} valid frame(s) "
                  f"({n_valid / len(idx):.1%}), held out by compound")
            for c in sorted(idx.loc[idx["split"] == "valid", "comp"].unique()):
                f = idx.loc[idx["comp"] == c, "formula"].iloc[0]
                print(f"[*]   valid: {c}  ({f}, {int((idx['comp'] == c).sum())} frames)")
            if n_valid == 0:
                print("[!]   nothing was held out -- too few compounds for this fraction")
            elif n_valid == len(idx):
                sys.exit("[ERROR] every compound went to validation; lower --valid-frac")

        splits = ([("train", idx[idx["split"] == "train"]),
                   ("valid", idx[idx["split"] == "valid"])]
                  if args.valid_frac > 0 else [("", idx)])

    idx_path = os.path.join(args.out, "dataset_index.csv")
    if args.dry_run:
        print(f"\n[dry-run] would write "
              + ", ".join(f"{len(g)} frame(s) to {os.path.join(args.out, n)}/"
                          for n, g in splits if len(g))
              + f" and the index to {idx_path}; nothing written.")
        return

    register_types(args.label_name)
    import dpdata

    os.makedirs(args.out, exist_ok=True)
    systems_of = {}
    for name, part in splits:
        if part.empty:
            continue
        dest = os.path.join(args.out, name) if name else args.out
        systems = []
        for path, g in part.groupby("path", sort=True):
            src = os.path.join(root, g.iloc[0]["structure"])
            s = dpdata.System(src, fmt="vasp/poscar", type_map=type_map)
            n = len(g)
            # Same cell repeated once per measurement: the structure is fixed and fparam
            # is what varies, which is the whole point of a frame parameter.
            s.data["coords"] = np.repeat(s.data["coords"], n, axis=0)
            s.data["cells"] = np.repeat(s.data["cells"], n, axis=0)
            t = convert_temperature(g["t_sheet"].to_numpy(float), sheet_unit,
                                    args.t_unit)
            p_hi, p_lo = g["p_high_atm"].to_numpy(float), g["p_low_atm"].to_numpy(float)
            if args.p_log10:
                p_hi, p_lo = np.log10(p_hi), np.log10(p_lo)
            s.data["fparam"] = np.stack([p_hi, p_lo, t], axis=1)
            s.data[args.label_name] = g[args.label_name].to_numpy(float).reshape(-1, 1)
            systems.append(s)

        ms = dpdata.MultiSystems(*systems, type_map=type_map)
        ms.to_deepmd_npy_mixed(dest, set_size=args.set_size)

        print(f"\n[*] wrote {len(part)} frame(s) to {dest}/ (deepmd/npy/mixed)")
        written = []
        for d in sorted(os.listdir(dest)):
            sub = os.path.join(dest, d)
            if not os.path.isdir(sub) or not d.isdigit():
                continue
            written.append(sub)
            nf = sum(np.load(os.path.join(sub, s, "coord.npy")).shape[0]
                     for s in sorted(os.listdir(sub)) if s.startswith("set."))
            print(f"[*]   {os.path.join(name, d) if name else d}/  {nf} frame(s), "
                  f"{d} atoms")
        systems_of[name] = written

    idx.to_csv(idx_path, index=False)
    if args.kfold:
        folds = fold_manifest(args.out, args.kfold,
                              {k: systems_of.get(f"fold_{k}", []) for k in range(args.kfold)})
        for k in range(args.kfold):
            folds[f"fold_{k}"]["valid_compounds"] = \
                sorted(idx.loc[idx["fold"] == k, "comp"].unique())
        folds_path = os.path.join(args.out, "folds.json")
        with open(folds_path, "w") as fh:
            json.dump(folds, fh, indent=2)
        print(f"\n[*] folds -> {folds_path}   (the systems lists for each of the "
              f"{args.kfold} runs)")
    print(f"\n[*] index -> {idx_path}   (a 'split' column records which side each frame "
          f"went to)")
    print(f"[*] sheet read as {sheet_unit}, fparam written in {args.t_unit}")
    print(f"[*] fparam is [P_high(atm), P_low(atm), T({args.t_unit})]"
          f"{' with pressures as log10' if args.p_log10 else ''}; "
          f"label file is {args.label_name}.npy")
    print(f"[*] set property_name = {args.label_name!r} and numb_fparam = 3 in the "
          f"deepmd input.json, with")
    if args.kfold:
        print(f"[*] each fold is written once; run {args.kfold} trainings, taking fold k's")
        print(f"[*]   validation_data.systems and training_data.systems from folds.json")
        print(f"[*] average the {args.kfold} validation scores, weighted by fold size; "
              f"every compound is held out exactly once")
    elif args.valid_frac > 0:
        print(f"[*]   training_data.systems   = {os.path.join(args.out, 'train')}/*")
        print(f"[*]   validation_data.systems = {os.path.join(args.out, 'valid')}/*")
    else:
        print(f"[*]   training_data.systems   = {args.out}/*")


if __name__ == "__main__":
    main()
