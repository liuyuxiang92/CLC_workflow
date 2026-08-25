#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stage 9c -- calculated oxygen swing vs the measured one, compound by compound.

    clc compare --calc dG0_vacancy_summary_old.csv \\
                                 --exp  dG0_vacancy_exp --outdir <dir>

WHAT IS COMPARED

  The experiment reports Δδ between 0.8 atm and 3e-5 atm.  δ is the oxygen
  non-stoichiometry per formula unit, and the calculation's x = n_v/n_fu is the
  same quantity, so the two are directly comparable with no conversion.

  **The x_swing column in dG0_vacancy_summary.csv is NOT that number.**
  build_simple_summary defines it against the highest pressure in --pO2-list,
  which is 1 atm, so the CSV's x_swing = x_3e-05atm - x_1atm.  The experiment
  swings from 0.8 atm.  This script therefore recomputes

      x_swing_0p8 = x_3e-05atm - x_0p8atm

  from the per-pressure columns, and reports the CSV's own x_swing beside it so
  the difference is visible rather than assumed away.

MATCHING

  Formulas are parsed into (Sr, Ca|Ba, Fe, Co) fractions and rebuilt as the
  directory-style key the calculation uses (Sr875_Ca125_Fe875_Co125,
  Sr1000_None000_Fe1000_None000, ...), so a match is exact rather than a
  nearest-neighbour guess.  Anything with no calculated counterpart -- the whole
  Sr/Ba family here -- is reported as unmatched instead of being dropped
  silently.

RELIABILITY

  The source workbook marks doubtful measurements with red text and says so in
  its header ("标红表示实验数据可信度较低").  That flag is read from the cell
  font and carried into the plots as open symbols, and every statistic is quoted
  both with and without those points.
"""

import argparse
import re
import shutil
import tempfile
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

RED_FONTS = {"FFC00000", "FFFF0000"}      # the workbook's "low confidence" reds


# ---------- experiment ----------

def _clean(s):
    """Strip zero-width joiners and normalise the minus sign in a formula."""
    s = unicodedata.normalize("NFKC", str(s))
    for ch in ("​", "‌", "‍", "﻿", " ", "\xa0"):
        s = s.replace(ch, "")
    return s.replace("−", "-")


_TOKEN = re.compile(r"([A-Z][a-z]?)(\d*\.?\d*)")


def parse_formula(s):
    """
    'Sr0.875Ca0.125Fe0.875Co0.125O3-d' -> dict of element -> fraction.

    A bare symbol means 1.0 (SrFeO3 is Sr 1.0, Fe 1.0).  O is ignored: the
    comparison is about the cation lattice.
    """
    out = {}
    for el, num in _TOKEN.findall(_clean(s)):
        if el == "O":
            continue
        out[el] = float(num) if num else 1.0
    return out


def comp_key(frac):
    """
    Rebuild the calculation's composition directory name from cation fractions.

    A-site is Sr + (Ca or Ba); B-site is Fe + Co.  An absent dopant is written
    None000, exactly as the calculated table spells it.
    """
    a_dop = "Ca" if "Ca" in frac else ("Ba" if "Ba" in frac else "None")
    b_dop = "Co" if "Co" in frac else "None"
    sr = frac.get("Sr", 0.0)
    fe = frac.get("Fe", 0.0)
    ad = frac.get(a_dop, 0.0) if a_dop != "None" else 0.0
    bd = frac.get(b_dop, 0.0) if b_dop != "None" else 0.0
    r = lambda v: int(round(v * 1000))
    return (f"Sr{r(sr):03d}_{a_dop}{r(ad):03d}_Fe{r(fe):03d}_{b_dop}{r(bd):03d}",
            a_dop, b_dop, sr, fe)


def read_exp(path):
    """
    Read the workbook into tidy rows: one per (compound, temperature).

    The sheet carries two header rows -- a merged 'Δδ (0.8-3e-5)' label over a
    670K/1070K pair -- so the temperatures are taken from the second row and the
    data starts below it.
    """
    src = Path(path)
    # openpyxl dispatches on the extension, and this file has none
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "exp.xlsx"
        shutil.copy(src, tmp)
        import openpyxl
        wb = openpyxl.load_workbook(tmp)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for r in range(3, ws.max_row + 1):          # row 1 note, row 2 header
            name = ws.cell(r, 2).value
            if not name or not str(name).strip():
                continue
            font = ws.cell(r, 2).font
            rgb = font.color.rgb if (font.color and font.color.type == "rgb") else None
            low_conf = str(rgb).upper() in RED_FONTS
            for col, T in ((3, 670), (4, 1070)):
                v = ws.cell(r, col).value
                if v is None:
                    continue
                rows.append({"exp_name": _clean(name), "Temperature_K": T,
                             "exp_swing": float(v), "low_confidence": low_conf})
    df = pd.DataFrame(rows)
    keys = df["exp_name"].map(lambda s: comp_key(parse_formula(s)))
    df["composition"] = [k[0] for k in keys]
    df["a_dopant"] = [k[1] for k in keys]
    df["b_dopant"] = [k[2] for k in keys]
    df["A_frac"] = [k[3] for k in keys]
    df["B_frac"] = [k[4] for k in keys]
    return df


# ---------- statistics ----------

def stats_block(x, y):
    """Agreement of y (calculated) with x (measured): correlation and error."""
    x = np.asarray(x, float); y = np.asarray(y, float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    n = x.size
    if n < 2:
        return {"n": int(n)}
    d = y - x
    out = {"n": int(n),
           "pearson_r": float(np.corrcoef(x, y)[0, 1]),
           "MAE": float(np.mean(np.abs(d))),
           "RMSE": float(np.sqrt(np.mean(d**2))),
           "bias_mean_calc_minus_exp": float(np.mean(d)),
           "exp_mean": float(np.mean(x)), "calc_mean": float(np.mean(y))}
    # Spearman without scipy: Pearson on the ranks
    rx = pd.Series(x).rank().to_numpy()
    ry = pd.Series(y).rank().to_numpy()
    out["spearman_rho"] = float(np.corrcoef(rx, ry)[0, 1])
    if np.ptp(x) > 0:
        m_, b_ = np.polyfit(x, y, 1)
        out["fit_slope"], out["fit_intercept"] = float(m_), float(b_)
    return out


# ---------- plot ----------
VARIANT_COLORS = ["#1f77b4", "#d62728", "#2ca02c", "#9467bd"]
# distinct shapes as well as colours: two variants that happen to be the same
# data land on exactly the same spot, and identical markers would hide that
VARIANT_MARKERS = ["o", "s", "^", "D"]


def plot_parity(df, temps, out_png, dpi=200, annotate=True, calc_on_x=False):
    """
    Parity panels, one per temperature.

    With a single calculation the points are coloured by A-site family; with
    several they are coloured by variant instead, because the question then is
    which variant tracks the measurement, not which family does.

    calc_on_x swaps the axes to x = calculated, y = measured.  Only the display
    changes: r, ρ, MAE and RMSE are symmetric, and `bias` stays defined as
    calculated - measured so it means the same thing in both orientations.
    """
    fig, axes = plt.subplots(1, len(temps), figsize=(6.2 * len(temps), 6.0))
    axes = np.atleast_1d(axes)
    fam_color = {"Ca": "#1f77b4", "None": "#d62728", "Ba": "#7f7f7f"}
    fam_label = {"Ca": "Sr/Ca B-site Fe(/Co)", "None": "Sr only (no A dopant)",
                 "Ba": "Sr/Ba (no calculation)"}
    variants = list(dict.fromkeys(df["variant"]))
    multi = len(variants) > 1
    var_color = {v: VARIANT_COLORS[i % len(VARIANT_COLORS)]
                 for i, v in enumerate(variants)}
    var_marker = {v: VARIANT_MARKERS[i % len(VARIANT_MARKERS)]
                  for i, v in enumerate(variants)}
    # sizes step down so a later variant sitting on an earlier one stays visible
    var_size = {v: 130 - 30 * i for i, v in enumerate(variants)}

    xcol, ycol = ("calc_swing", "exp_swing") if calc_on_x else ("exp_swing", "calc_swing")
    lab = {"exp_swing": "measured Δδ  (0.8 → 3×10$^{-5}$ atm)",
           "calc_swing": "calculated Δx = $x_{3e-5}$ − $x_{0.8}$"}

    for ax, T in zip(axes, temps):
        d = df[(df["Temperature_K"] == T) & df["calc_swing"].notna()]
        lim_hi = max(1e-3, np.nanmax([d["exp_swing"].max(), d["calc_swing"].max()])) * 1.15
        ax.plot([0, lim_hi], [0, lim_hi], "k--", lw=1.0, alpha=0.7, label="y = x")

        group_col = "variant" if multi else "a_dopant"
        for key, g in d.groupby(group_col):
            c = var_color[key] if multi else fam_color.get(key, "#555555")
            name = key if multi else fam_label.get(key, key)
            mk = var_marker[key] if multi else "o"
            sz = var_size[key] if multi else 70
            for low, gg in g.groupby("low_confidence"):
                ax.scatter(gg[xcol], gg[ycol], s=sz, marker=mk,
                           facecolors="none" if low else c, edgecolors=c,
                           linewidths=1.6, alpha=0.85,
                           label=f"{name}"
                                 f"{' (low-confidence exp)' if low else ''}")
        if annotate and not multi:
            # only the ends of the measured range: the calculated values sit in a
            # band so narrow that labelling every point buries the panel in
            # overlapping text.  The ranked bar chart carries all the names.
            dd = d.sort_values("exp_swing")
            for _, r in pd.concat([dd.head(3), dd.tail(3)]).iterrows():
                ax.annotate(r["short"], (r[xcol], r[ycol]),
                            textcoords="offset points", xytext=(6, 5), fontsize=7,
                            color="#333333")

        lines = []
        for v in variants:
            dv = d[d["variant"] == v]
            s = stats_block(dv["exp_swing"], dv["calc_swing"])
            s_hi = stats_block(dv.loc[~dv["low_confidence"], "exp_swing"],
                               dv.loc[~dv["low_confidence"], "calc_swing"])
            head = f"{v}:  " if multi else ""
            lines.append(f"{head}n = {s.get('n', 0)}   "
                         f"r = {s.get('pearson_r', np.nan):+.2f}   "
                         f"ρ = {s.get('spearman_rho', np.nan):+.2f}")
            lines.append(f"    MAE = {s.get('MAE', np.nan):.3f}   "
                         f"bias = {s.get('bias_mean_calc_minus_exp', np.nan):+.3f}   "
                         f"[high-conf r = {s_hi.get('pearson_r', np.nan):+.2f}]")
        txt = "\n".join(lines)
        ax.text(0.03, 0.97, txt, transform=ax.transAxes, va="top", fontsize=8.5,
                bbox=dict(boxstyle="round,pad=0.4", fc="white", ec="#cccccc", alpha=0.9))

        ax.set_xlim(0, lim_hi); ax.set_ylim(0, lim_hi)
        ax.set_aspect("equal", adjustable="box")
        ax.set_xlabel(lab[xcol], fontsize=10)
        ax.set_ylabel(lab[ycol], fontsize=10)
        ax.set_title(f"{T} K", fontsize=12)
        ax.grid(alpha=0.25, lw=0.6)
        h, l = ax.get_legend_handles_labels()
        seen, hh, ll = set(), [], []
        for a, b in zip(h, l):
            if b not in seen:
                seen.add(b); hh.append(a); ll.append(b)
        ax.legend(hh, ll, fontsize=7.5, loc="lower right", framealpha=0.9)

    fig.suptitle("Oxygen swing: calculation vs experiment, matched compounds\n"
                 "open symbols = measurements the source workbook flags as "
                 "low-confidence", fontsize=12)
    fig.tight_layout(rect=[0, 0, 1, 0.92])
    fig.savefig(out_png, dpi=dpi)
    plt.close(fig)
    print(f"[*] wrote {out_png}")


def plot_ranked(df, temps, out_png, dpi=200):
    """Per-compound side-by-side bars, sorted by the measured value."""
    fig, axes = plt.subplots(len(temps), 1, figsize=(11, 4.2 * len(temps)))
    axes = np.atleast_1d(axes)
    variants = list(dict.fromkeys(df["variant"]))
    for ax, T in zip(axes, temps):
        sub = df[(df["Temperature_K"] == T) & df["calc_swing"].notna()]
        # one row per compound, ordered by the measurement; the variants ride along
        order = (sub[sub["variant"] == variants[0]]
                 .sort_values("exp_swing")["short"].tolist())
        i = np.arange(len(order))
        n_bar = len(variants) + 1
        w = 0.8 / n_bar
        base = sub[sub["variant"] == variants[0]].set_index("short").loc[order]
        ax.bar(i - 0.4 + w / 2, base["exp_swing"], w,
               label="experiment", color="#4c72b0")
        for j, v in enumerate(variants):
            g = sub[sub["variant"] == v].set_index("short").reindex(order)
            ax.bar(i - 0.4 + w * (j + 1.5), g["calc_swing"], w,
                   label=f"calculation ({v})" if len(variants) > 1 else "calculation",
                   color=VARIANT_COLORS[(j + 1) % len(VARIANT_COLORS)])
        for k, low in enumerate(base["low_confidence"]):
            if low:
                ax.text(i[k] - 0.4 + w / 2, base["exp_swing"].iloc[k], "*", ha="center",
                        va="bottom", fontsize=12, color="#b03030")
        ax.set_xticks(i)
        ax.set_xticklabels(order, rotation=45, ha="right", fontsize=7.5)
        ax.set_ylabel("Δδ  (0.8 → 3e-5 atm)", fontsize=10)
        ax.set_title(f"{T} K   (* = low-confidence measurement)", fontsize=11)
        ax.grid(axis="y", alpha=0.25, lw=0.6)
        ax.legend(fontsize=9)
    fig.tight_layout()
    fig.savefig(out_png, dpi=dpi)
    plt.close(fig)
    print(f"[*] wrote {out_png}")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--calc", required=True, nargs="+",
                    help="one or more dG0_vacancy_summary*.csv; give several to "
                         "overlay variants (e.g. |a|=2 vs |a|=0) on one figure")
    ap.add_argument("--labels", default=None,
                    help="comma-separated names for the --calc files, in order "
                         "(default: their file stems)")
    ap.add_argument("--exp", required=True, help="the experimental workbook")
    ap.add_argument("--outdir", default=".", help="where CSV and PNGs go")
    ap.add_argument("--no-annotate", action="store_true",
                    help="do not label each point on the parity plot")
    ap.add_argument("--calc-on-x", action="store_true",
                    help="plot x = calculated, y = measured (default is the "
                         "other way round). Display only: every statistic is "
                         "unchanged, and bias stays calculated - measured.")
    ap.add_argument("--dpi", type=int, default=200)
    args = ap.parse_args(argv)

    outdir = Path(args.outdir).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    labels = ([s.strip() for s in args.labels.split(",")] if args.labels
              else [Path(p).stem for p in args.calc])
    if len(labels) != len(args.calc):
        raise SystemExit(f"--labels has {len(labels)} name(s) for "
                         f"{len(args.calc)} --calc file(s)")

    exp = read_exp(args.exp)
    print(f"[*] experiment: {exp['exp_name'].nunique()} compound(s), "
          f"{len(exp)} (compound, T) measurement(s)")
    print(f"[*] {exp.loc[exp['low_confidence'], 'exp_name'].nunique()} compound(s) "
          f"flagged low-confidence by the workbook")

    frames = []
    for path, label in zip(args.calc, labels):
        calc = pd.read_csv(path)
        # the number the experiment actually measures: 0.8 atm -> 3e-5 atm
        calc["calc_swing"] = (calc["x_3e-05atm"] - calc["x_0p8atm"]).round(6)
        calc = calc.rename(columns={"x_swing": "calc_swing_vs_1atm"})
        keep = ["composition", "Temperature_K", "calc_swing", "calc_swing_vs_1atm",
                "x_1atm", "x_0p8atm", "x_3e-05atm", "R2", "n_sets_used",
                "n_sets_total", "family"]
        keep = [c for c in keep if c in calc.columns]
        m = exp.merge(calc[keep], on=["composition", "Temperature_K"], how="left")
        m.insert(0, "variant", label)
        frames.append(m)
    merged = pd.concat(frames, ignore_index=True)

    # a short label for the plots
    merged["short"] = (merged["exp_name"].str.replace("O3-d", "", regex=False)
                                          .str.replace("O3-δ", "", regex=False))

    matched = merged[merged["calc_swing"].notna()]
    unmatched = merged[merged["calc_swing"].isna()]
    n_m = matched["composition"].nunique()
    n_u = unmatched["composition"].nunique()
    print(f"[*] matched {n_m} compound(s) to the calculation; {n_u} unmatched")
    if n_u:
        for c in sorted(unmatched["exp_name"].unique()):
            print(f"      no calculation for {c}")
    print(f"[*] variants: {', '.join(labels)}")

    merged["residual_calc_minus_exp"] = merged["calc_swing"] - merged["exp_swing"]
    p = outdir / "exp_vs_calc_swing.csv"
    merged.to_csv(p, index=False)
    print(f"[*] wrote {p}")

    temps = sorted(matched["Temperature_K"].unique())
    plot_parity(merged, temps, outdir / "exp_vs_calc_parity.png",
                dpi=args.dpi, annotate=not args.no_annotate,
                calc_on_x=args.calc_on_x)
    plot_ranked(merged, temps, outdir / "exp_vs_calc_ranked.png", dpi=args.dpi)

    print("\n=== agreement (calculated vs measured Δδ, 0.8 -> 3e-5 atm) ===")
    for T in temps:
        for v in labels:
            d = matched[(matched["Temperature_K"] == T) & (matched["variant"] == v)]
            for tag, dd in (("all matched", d),
                            ("high-confidence only", d[~d["low_confidence"]])):
                s = stats_block(dd["exp_swing"], dd["calc_swing"])
                if s.get("n", 0) < 2:
                    continue
                print(f"  {T} K  {v:<28s} {tag:22s} n={s['n']:2d}  "
                      f"r={s['pearson_r']:+.3f}  rho={s['spearman_rho']:+.3f}  "
                      f"MAE={s['MAE']:.4f}  RMSE={s['RMSE']:.4f}  "
                      f"bias={s['bias_mean_calc_minus_exp']:+.4f}  "
                      f"(exp mean {s['exp_mean']:.4f}, calc mean {s['calc_mean']:.4f})")


if __name__ == "__main__":
    main()
